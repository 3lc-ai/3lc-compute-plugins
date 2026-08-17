# Copyright 2026 3LC Inc.
# SPDX-License-Identifier: Apache-2.0
"""Export plugin — export 3LC tables to YOLO, COCO, CSV, and XLSX formats."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

from tlc_plugin_sdk import ComputePlugin

if TYPE_CHECKING:
    from tlc_plugin_sdk import JobContext

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Export format definitions
# ---------------------------------------------------------------------------

EXPORT_FORMATS: dict[str, dict[str, Any]] = {
    "yolo": {
        "name": "yolo",
        "display_name": "YOLO Format",
        "description": "Export table to YOLO dataset format (labels + images).",
        "icon": "⊡",
        "file_extension": ".yaml",
        "options": [
            {
                "name": "split",
                "label": "Split name",
                "type": "select",
                "choices": ["train", "val", "test"],
                "default": "train",
                "help": "Which split this table represents in the YOLO dataset.",
            },
            {
                "name": "image_strategy",
                "label": "Image handling",
                "type": "select",
                "choices": ["ignore", "copy", "symlink"],
                "default": "copy",
                "help": '"ignore" exports labels only, "copy" copies images alongside labels, '
                '"symlink" creates symlinks (not supported on Windows).',
            },
            {
                "name": "weight_threshold",
                "label": "Weight threshold",
                "type": "text",
                "default": "0.0",
                "help": "Exclude rows with weight below this value (0.0 keeps all rows).",
            },
        ],
    },
    "coco": {
        "name": "coco",
        "display_name": "COCO Format",
        "description": "Export table to COCO JSON annotations format.",
        "icon": "⊡",
        "file_extension": ".json",
        "options": [
            {
                "name": "weight_threshold",
                "label": "Weight threshold",
                "type": "text",
                "default": "0.0",
                "help": "Exclude rows with weight below this value (0.0 keeps all rows).",
            },
        ],
    },
    "csv": {
        "name": "csv",
        "display_name": "CSV",
        "description": "Export table columns to CSV spreadsheet. Complex types are serialized as JSON strings; "
        "URL aliases are expanded to absolute paths.",
        "icon": "⊞",
        "file_extension": ".csv",
        "needs_column_select": True,
        "options": [
            {
                "name": "weight_threshold",
                "label": "Weight threshold",
                "type": "text",
                "default": "0.0",
                "help": "Exclude rows with weight below this value (0.0 keeps all rows).",
            },
            {
                "name": "separator",
                "label": "Separator",
                "type": "select",
                "choices": [",", ";", "tab"],
                "default": ",",
                "help": "Column separator character.",
            },
        ],
    },
    "xlsx": {
        "name": "xlsx",
        "display_name": "Excel (XLSX)",
        "description": "Export table columns to Excel spreadsheet. Complex types are serialized as JSON strings; "
        "URL aliases are expanded to absolute paths.",
        "icon": "⊞",
        "file_extension": ".xlsx",
        "needs_column_select": True,
        "options": [
            {
                "name": "weight_threshold",
                "label": "Weight threshold",
                "type": "text",
                "default": "0.0",
                "help": "Exclude rows with weight below this value (0.0 keeps all rows).",
            },
        ],
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _serialize_value(value: Any, *, is_url: bool = False) -> Any:
    """Serialize a cell value for CSV/XLSX export.

    Simple types pass through. Complex types (dict, list) become JSON strings.
    Bytes are hex-encoded. URL-column values (``is_url``) have alias tokens
    like ``<PYRO_IMAGES>/a.jpg`` expanded to absolute paths — alias form is
    3LC-internal representation and should not leak into exported files.
    """
    if value is None:
        return ""
    if is_url and isinstance(value, str) and value:
        import tlc

        return tlc.Url(value).expand_aliases(allow_unexpanded=True).to_str()
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, default=str, ensure_ascii=False)
    return str(value)


def _classify_column_type(schema: Any, col_name: str = "") -> str:
    """Classify a column schema into a human-readable type string."""
    try:
        schema_json = json.loads(schema.to_json()) if hasattr(schema, "to_json") else {}
    except Exception:
        schema_json = {}

    # Flatten: schema has value.type for simple, values.* for complex
    full_str = json.dumps(schema_json).lower()
    value = schema_json.get("value", {})
    value_type = str(value.get("type", "")).lower() if isinstance(value, dict) else ""
    string_role = str(value.get("string_role", "")).lower() if isinstance(value, dict) else ""
    has_values = "values" in schema_json  # complex nested type (bbs, segmentation)

    # Image columns (string_role contains "image")
    if "image" in string_role:
        return "image"
    # Bounding boxes (nested with bb_list)
    if has_values and ("bb_list" in full_str or "bounding_box" in full_str):
        return "bbox"
    # Segmentation (nested with rles or segmentation)
    if has_values and ("rle" in full_str or "segmentation" in full_str):
        return "segmentation"
    # Embeddings (by column name convention)
    if "embedding" in col_name.lower():
        return "embedding"
    # Weight column
    if "sample_weight" in full_str or col_name == "weight":
        return "number"
    # Simple numeric types
    if value_type in ("int32", "int64", "float32", "float64", "double"):
        return "number"
    # String
    if value_type == "string":
        return "string"
    # Boolean
    if value_type in ("bool", "boolean"):
        return "boolean"
    # Complex nested (has values dict but didn't match above)
    if has_values:
        return "complex"
    return "other"


# ---------------------------------------------------------------------------
# Execution logic
# ---------------------------------------------------------------------------


def _execute_yolo(table_url: str, output_path: str, options: dict[str, Any]) -> dict[str, Any]:
    """Execute YOLO export."""
    import tlc

    split = options.get("split", "train") or "train"
    image_strategy = options.get("image_strategy", "copy") or "copy"
    weight_threshold = float(options.get("weight_threshold", "0") or "0")

    table = tlc.Table.from_url(table_url)
    table.export(
        output_url=output_path,
        format="yolo",
        split=split,
        image_strategy=image_strategy,
        weight_threshold=weight_threshold,
    )
    return {
        "success": True,
        "message": f"Exported to YOLO format at {output_path}",
        "details": {
            "output_path": output_path,
            "split": split,
            "image_strategy": image_strategy,
            "row_count": table.row_count,
        },
    }


def _execute_coco(table_url: str, output_path: str, options: dict[str, Any]) -> dict[str, Any]:
    """Execute COCO export."""
    import tlc

    p = Path(output_path)
    if p.suffix.lower() == ".json":
        out = str(p)
    elif p.is_dir():
        out = str(p / "annotations.json")
    else:
        p.mkdir(parents=True, exist_ok=True)
        out = str(p / "annotations.json")

    weight_threshold = float(options.get("weight_threshold", "0") or "0")

    table = tlc.Table.from_url(table_url)
    table.export(output_url=out, format="coco", weight_threshold=weight_threshold)
    return {
        "success": True,
        "message": f"Exported to COCO JSON at {out}",
        "details": {"output_path": out, "row_count": table.row_count},
    }


def _execute_csv(table_url: str, output_path: str, options: dict[str, Any]) -> dict[str, Any]:
    """Execute CSV export with optional column selection."""
    import tlc

    weight_threshold = float(options.get("weight_threshold", "0") or "0")
    separator = options.get("separator", ",") or ","
    if separator == "tab":
        separator = "\t"
    selected_columns: list[str] | None = options.get("columns")

    table = tlc.Table.from_url(table_url)

    # Determine output file path
    p = Path(output_path)
    if p.suffix.lower() != ".csv":
        p.mkdir(parents=True, exist_ok=True)
        p = p / "export.csv"
    else:
        p.parent.mkdir(parents=True, exist_ok=True)

    # Get all column names from first row
    all_columns = list(table.table_rows[0].keys())
    columns = selected_columns or all_columns

    from tlc_plugin_sdk.shared.url_utils import get_url_column_names

    url_columns = set(get_url_column_names(table))

    # Write CSV
    row_count = 0
    with open(p, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=columns,
            delimiter=separator,
            extrasaction="ignore",
            lineterminator="\n",
        )
        writer.writeheader()

        weights_col = table.weights_column_name if hasattr(table, "weights_column_name") else None
        for row in table.table_rows:
            if weight_threshold > 0 and weights_col and cast("float", row.get(weights_col, 1.0)) < weight_threshold:
                continue
            writer.writerow({col: _serialize_value(row.get(col), is_url=col in url_columns) for col in columns})
            row_count += 1

    return {
        "success": True,
        "message": f"Exported {row_count} rows to CSV at {p}",
        "details": {
            "output_path": str(p),
            "row_count": row_count,
            "columns": len(columns),
        },
    }


def _execute_xlsx(table_url: str, output_path: str, options: dict[str, Any]) -> dict[str, Any]:
    """Execute XLSX export with optional column selection."""
    import tlc

    weight_threshold = float(options.get("weight_threshold", "0") or "0")
    selected_columns: list[str] | None = options.get("columns")

    table = tlc.Table.from_url(table_url)

    # Determine output file path
    p = Path(output_path)
    if p.suffix.lower() != ".xlsx":
        p.mkdir(parents=True, exist_ok=True)
        p = p / "export.xlsx"
    else:
        p.parent.mkdir(parents=True, exist_ok=True)

    # Get all column names from first row
    all_columns = list(table.table_rows[0].keys())
    columns = selected_columns or all_columns

    from tlc_plugin_sdk.shared.url_utils import get_url_column_names

    url_columns = set(get_url_column_names(table))

    # Collect rows
    rows: list[dict[str, Any]] = []
    weights_col = table.weights_column_name if hasattr(table, "weights_column_name") else None
    for row in table.table_rows:
        if weight_threshold > 0 and weights_col and cast("float", row.get(weights_col, 1.0)) < weight_threshold:
            continue
        rows.append({col: _serialize_value(row.get(col), is_url=col in url_columns) for col in columns})

    # Write XLSX using openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            # openpyxl doesn't handle very long strings well — truncate if needed
            if isinstance(value, str) and len(value) > 32767:
                value = value[:32767]
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(p))

    return {
        "success": True,
        "message": f"Exported {len(rows)} rows to Excel at {p}",
        "details": {
            "output_path": str(p),
            "row_count": len(rows),
            "columns": len(columns),
        },
    }


_EXECUTORS: dict[str, Any] = {
    "yolo": _execute_yolo,
    "coco": _execute_coco,
    "csv": _execute_csv,
    "xlsx": _execute_xlsx,
}


# ---------------------------------------------------------------------------
# Plugin class
# ---------------------------------------------------------------------------


class ExportPlugin(ComputePlugin):
    """Sidebar plugin for exporting 3LC tables to external formats.

    Behavior only — all metadata lives in ``plugin.toml`` (the manifest). The host
    instantiates this via the manifest's ``runtime.entrypoint`` and stamps ``id``
    onto the instance; the class does not declare it.
    """

    _ui_cache: str | None = None

    def get_ui_fragment(self) -> str:
        """Return the self-contained export wizard HTML+JS+CSS fragment."""
        if self._ui_cache is None:
            from tlc_plugin_sdk.shared.alias_override_ui import alias_override_ui_script
            from tlc_plugin_sdk.shared.data_source_ui import data_source_ui_script
            from tlc_plugin_sdk.shared.job_tracker import job_tracker_script
            from tlc_plugin_sdk.shared.ui_inject import inject_scripts

            ui_path = Path(__file__).resolve().parent / "ui.html"
            raw = ui_path.read_text(encoding="utf-8")
            self._ui_cache = inject_scripts(raw, data_source_ui_script(), alias_override_ui_script(), job_tracker_script())
        return self._ui_cache

    def compute(self, params: dict[str, Any]) -> dict[str, Any]:
        """Return export format definitions."""
        return {"formats": list(EXPORT_FORMATS.values())}

    def run_job(self, ctx: JobContext) -> None:
        """Run one export as a host-managed job.

        Long exports (a YOLO export copies every image and can run for minutes)
        must not ride a single synchronous request: the Hub's fetch gives up at
        its timeout and reports failure for work that then completes anyway. On
        the job channel, progress reaches the generic Queue & Progress panel,
        completion is a broadcast instead of a held connection, and the rich
        result reaches the fragment as an ``export_result`` event.

        Args:
            ctx: Host-provided job context. ``ctx.params`` carries the same body
                the ``/execute`` route accepts: ``format``, ``table_url``,
                ``output_path``, format-specific options, and optional
                ``alias_overrides``.

        Raises:
            ValueError: When the export fails, so the host marks the job failed
                and the message reaches the UI via the generic channel.

        """
        from tlc_plugin_exporter.routes import run_export

        format_name = str(ctx.params.get("format", "") or "")
        label = f"Exporting to {format_name.upper()}…" if format_name else "Exporting…"
        ctx.progress(percent=-1, label=label)

        result = run_export(ctx.params)

        payload = dict(result)
        payload["job_id"] = ctx.job_id
        ctx.emit("export_result", payload)

        if not result.get("success"):
            raise ValueError(str(result.get("message") or "Export failed"))
        ctx.progress(percent=100, label="Export complete")

    def get_route_handlers(self) -> list[Any]:
        """Serve the export plugin's custom routes as relative Litestar handlers."""
        from tlc_plugin_sdk.shared.data_source_routes import data_source_route_handlers

        from tlc_plugin_exporter import routes as _routes

        return [*_routes.get_route_handlers(), *data_source_route_handlers()]
